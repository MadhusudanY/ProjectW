import os
import re
from pathlib import Path

import docx2txt
from openpyxl import Workbook

from PyPDF2 import PdfReader


# Function to get folder containing position descriptions documents from the console input
def get_path():
    s = input("Please enter folder path : ")
    c = input("Please enter customer name : ").upper()
    return s.replace('\\', '/'), c


def get_headers(client_name):
    headers_dict = {}
    with open('Headers.txt', 'r') as f:
        for line in f.readlines():
            if line.startswith(client_name):
                headers_list = line.split('|')
                for x in headers_list[1:]:
                    if x.count('-') == 0:
                        headers_dict[x] = ""
                    else:
                        x = x.split("-")[0]
                        headers_dict[x] = ""
                f.close()
                break
    return headers_dict


def get_chug_it(client_name):
    chug_it = 0
    with open('Headers.txt', 'r') as f:
        for line in f.readlines():
            if line.startswith(client_name):
                chug_it = int(line.split('|')[-1])
                f.close()
                break
    return chug_it * -1


def get_col_data(client, headers, data):
    some_list = list(headers.keys())
    some_list.remove('FILENAME')
    marker1, marker2 = "", ""
    for i in range(len(some_list) - 1):
        marker1 = some_list[i]
        marker2 = some_list[i + 1]
        col_data = ""
        regex_pattern = re.compile(f'^{marker1}([\S\s]*)^{marker2}', re.MULTILINE)
        if regex_pattern.search(data) is not None:
            col_data = regex_pattern.search(data).group().replace(marker1, "").replace(marker2, "").strip()
        elif regex_pattern.search(data) is None:
            if data.lower().count(marker1.lower()) > 0 and data.lower().count(marker2.lower()) > 0:
                start_index = data.lower().index(marker1.lower()) + len(marker1)
                end_index = data.lower().index(marker2.lower())
                col_data = data[start_index:end_index].strip()
            elif data.count(marker1) > 0 and data.count(marker2) > 0:
                start_index = data.index(marker1) + len(marker1)
                end_index = data.index(marker2)
                col_data = data[start_index:end_index].strip()
        headers[marker1] = col_data
    if data.lower().count(marker2.lower()) > 0:
        start_index = data.lower().index(marker2.lower()) + len(marker2)
        chug_it = get_chug_it(client)
        if chug_it < 0:
            col_data = data[start_index:chug_it].strip()
        else:
            col_data = data[start_index:].strip()
        headers[marker2] = col_data
    return headers


def modify_operation(col_header, modify_item, dict_headers):
    value = dict_headers.get(col_header)
    if modify_item.startswith('replaceWithEmpty'):
        a, b = modify_item.split(' ')
        value = value.replace(b, '')
        dict_headers[col_header] = value
    if modify_item.startswith('sliceLastIndexOf'):
        a, b, c = modify_item.split(' ')
        c = int(c)
        if c == 0:
            c = -1
        value = value[len(value) - value[::-1].index(b):]
        dict_headers[col_header] = value
    if modify_item.startswith('replaceAsContains'):
        a, b, c = modify_item.split(' ')
        if value.lower().__contains__(b.lower()):
            value = b
        elif value.lower().__contains__(c.lower()):
            value = c
        dict_headers[col_header] = value
    if modify_item.startswith('replaceFromKey'):
        key = modify_item[modify_item.index('(') + 1:modify_item.index(')')]
        value = dict_headers.get(key)
        a, b = modify_item.replace(f'replaceFromKey({key}) ', '').split('-')
        if a == 'AlphaNumericOneChar':
            value = re.findall("[a-zA-Z0-9]", value)[int(b)]
        dict_headers[col_header] = value
    if modify_item.startswith('BreakBacK'):
        key_list = list(dict_headers)
        key_index = key_list.index(col_header)
        values_reverse = value.split('\n\n')[::-1]
        for x in values_reverse:
            dict_headers[key_list[key_index]] = x
            key_index = key_index - 1
    if modify_item.startswith('LTrim'):
        dict_headers[col_header] = value.lstrip()
    return dict_headers


def apply_modifiers(client, dict_headers):
    modifiers_list = []
    with open('Modifiers.txt', 'r') as f:
        for line in f.readlines():
            if line.startswith(client):
                modifiers_list = line.split('|')[1:]
                f.close()
                break
    for item in modifiers_list:
        key, mod = item.split('#')
        dict_headers = modify_operation(key, mod, dict_headers)
    return dict_headers


def begin_process(client, text_data, filename, ws, row):
    # This is for debugging Purpose
    # print(text_data)
    # exit(0)
    dict_headers = get_headers(client)
    dict_headers['FILENAME'] = filename
    # Last line Item on Header.txt is the string value that needs to be excluded. I called it chug_it value.
    # Hence, pop it from dict_headers
    dict_headers.popitem()
    dict_headers = get_col_data(client, dict_headers, text_data)
    # This is for debugging Purpose
    # print(dict_headers)
    dict_headers = apply_modifiers(client, dict_headers)
    col = 1
    for k in dict_headers.keys():
        ws.cell(1, col).value = k.replace(":", "")
        ws.cell(row, col).value = dict_headers.get(k)
        col += 1
    row += 1
    return ws, row


def execute():
    pd_dir, client = get_path()
    wb = Workbook()
    xl_file = f"{pd_dir}/PositionDescriptions.xlsx"
    ws = wb.create_sheet(index=0, title="PositionDescriptions")
    row = 2
    try:
        for file in os.listdir(pd_dir):
            if file.endswith('.docx'):
                word_data = docx2txt.process(pd_dir+'\\' + file)
                try:
                    ws, row = begin_process(client, word_data, file, ws, row)
                except IndexError:
                    print("Something went wrong with " + file)
                    continue
            if file.endswith('.pdf'):
                with open(pd_dir+'\\' + file, mode='rb') as x:
                    reader = PdfReader(x)
                    pdf_data = ""
                    for i in range(len(reader.pages)):
                        page = reader.pages[i]
                        pdf_data = pdf_data + page.extract_text()
                    x.close()
                # print(pdf_data)
                ws, row = begin_process(client, pdf_data, file, ws, row)
    except PermissionError as p:
        print("Permission denied. Please see the docx file(s) are closed." + p.strerror)
    finally:
        wb.save(xl_file)
        print("Execution Complete")
